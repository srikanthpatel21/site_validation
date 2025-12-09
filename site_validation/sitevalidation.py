import requests
from openpyxl import load_workbook
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_FILE = os.path.join(BASE_DIR, "sites.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "sites_report.xlsx")

def check_url_status(url):
    try:
        response = requests.get(url, timeout=5)
        return response.status_code
    except requests.exceptions.RequestException:
        return "Error"

def process_excel():
    wb = load_workbook(INPUT_FILE)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row=row, column=1).value

        if not url:
            continue

        status = check_url_status(url)
        sheet.cell(row=row, column=2).value = status
        print(f"{url} → {status}")

    wb.save(OUTPUT_FILE)
    print("\n✔ Report generated:", OUTPUT_FILE)

if __name__ == "__main__":
    process_excel()
